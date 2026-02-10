"""
Milheiro → Unidade Converter
Backend API - Preserva formatação original do Excel.
"""

from flask import Flask, request, jsonify, send_from_directory, send_file
import pandas as pd
import numpy as np
import openpyxl
from copy import copy
import os
import re
import json
import sys
import webbrowser
import threading

# Suporte PyInstaller: quando empacotado, recursos ficam em _MEIPASS
if getattr(sys, 'frozen', False):
    # Executando como .exe
    BUNDLE_DIR = sys._MEIPASS
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = BUNDLE_DIR

# Static folder: prioriza bundle (PyInstaller), depois pasta local
STATIC_CANDIDATES = [
    os.path.join(BUNDLE_DIR, 'static'),
    os.path.join(BASE_DIR, 'static'),
    BUNDLE_DIR,
    BASE_DIR,
]
STATIC_FOLDER = next((p for p in STATIC_CANDIDATES if os.path.isfile(os.path.join(p, 'index.html'))), BASE_DIR)

app = Flask(__name__, static_folder=STATIC_FOLDER)

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ─── Helpers ──────────────────────────────────────────────

def clean_numeric_value(val):
    """
    Converte qualquer valor para número.
    Lida com: "R$ 1.234,56", "1.234,56", "1234.56", "1,5", "850", etc.
    """
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        if isinstance(val, float) and np.isnan(val):
            return np.nan
        return float(val)

    if not isinstance(val, str):
        val = str(val)

    val = val.strip()
    if val == '' or val == '-':
        return np.nan

    # Remove R$, US$, €, espaços
    val = re.sub(r'[R$€£US\s]', '', val)
    val = val.strip()
    if val == '':
        return np.nan

    has_comma = ',' in val
    has_dot = '.' in val

    if has_comma and has_dot:
        last_comma = val.rfind(',')
        last_dot = val.rfind('.')
        if last_comma > last_dot:
            val = val.replace('.', '').replace(',', '.')
        else:
            val = val.replace(',', '')
    elif has_comma and not has_dot:
        parts = val.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2:
            val = val.replace(',', '.')
        else:
            val = val.replace(',', '')
    elif has_dot and not has_comma:
        parts = val.split('.')
        if len(parts) == 2 and len(parts[1]) == 3 and len(parts[0]) <= 3:
            val = val.replace('.', '')
        elif len(parts) > 2:
            val = val.replace('.', '')

    try:
        return float(val)
    except ValueError:
        return np.nan


def try_convert_column_to_numeric(series):
    """Tenta converter uma série para numérico e retorna a taxa de conversão."""
    converted = series.apply(clean_numeric_value)
    non_null_original = series.dropna()
    non_null_original = non_null_original[non_null_original.astype(str).str.strip() != '']
    non_null_converted = converted.dropna()

    if len(non_null_original) == 0:
        return converted, 0.0

    ratio = len(non_null_converted) / len(non_null_original)
    return converted, ratio


def detect_header_row(df_raw):
    """Detecta a linha real de cabeçalho na planilha."""
    if df_raw.empty:
        return 0

    best_header_row = 0
    best_score = -1
    num_cols = df_raw.shape[1]
    max_check = min(15, len(df_raw))

    for i in range(max_check):
        row = df_raw.iloc[i]
        score = 0
        filled = 0

        for val in row:
            if pd.isna(val) or str(val).strip() == '':
                continue
            filled += 1
            sval = str(val).strip()
            is_pure_number = bool(re.match(r'^[\d.,R$€£\s\-+%]+$', sval))
            is_short = len(sval) < 60

            if not is_pure_number and is_short:
                score += 2
            elif not is_pure_number:
                score += 1

        if filled > 0 and filled >= max(2, num_cols * 0.3):
            normalized_score = (score / filled) * (filled / num_cols)
        else:
            normalized_score = 0

        if i + 1 < len(df_raw):
            next_row = df_raw.iloc[i + 1]
            next_has_numbers = sum(
                1 for v in next_row if not pd.isna(v) and
                (isinstance(v, (int, float)) or clean_numeric_value(v) is not np.nan)
            )
            if next_has_numbers > 0:
                normalized_score *= 1.5

        if normalized_score > best_score:
            best_score = normalized_score
            best_header_row = i

    return best_header_row


def analyze_columns(df):
    """Analisa colunas e retorna info sobre cada uma."""
    columns_info = []
    for col in df.columns:
        col_data = df[col]
        if pd.api.types.is_numeric_dtype(col_data):
            non_null = col_data.dropna()
            columns_info.append({
                'name': str(col),
                'type': 'numeric',
                'convertible': True,
                'numeric_ratio': 1.0,
                'sample_values': [format_sample(v) for v in non_null.head(5).tolist()],
                'non_null_count': len(non_null)
            })
        else:
            converted, ratio = try_convert_column_to_numeric(col_data)
            non_null_original = col_data.dropna()
            columns_info.append({
                'name': str(col),
                'type': 'text' if ratio < 0.5 else 'text_numeric',
                'convertible': ratio >= 0.5,
                'numeric_ratio': round(ratio, 2),
                'sample_values': [str(v) for v in non_null_original.head(5).tolist()],
                'non_null_count': len(non_null_original)
            })
    return columns_info


def format_sample(v):
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else f"{v:.2f}"
    return str(v)


def sanitize_preview(preview):
    for row in preview:
        for key in row:
            if isinstance(row[key], (np.integer,)):
                row[key] = int(row[key])
            elif isinstance(row[key], (np.floating,)):
                row[key] = '' if np.isnan(row[key]) else float(row[key])
    return preview


def convert_excel_preserving_format(filepath, output_path, header_row, columns_to_convert, divisor):
    """
    Converte colunas no Excel preservando TODA a formatação original:
    - Bordas, fontes, cores, alinhamento
    - Largura de colunas, altura de linhas
    - Células mescladas
    - Cabeçalhos originais (todas as linhas acima dos dados)
    - Estilos condicionais
    
    Copia o arquivo original e altera apenas os valores das células selecionadas.
    """
    # Carrega o workbook preservando tudo
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Mapear nomes de colunas para índices (1-based)
    # A linha de header no openpyxl é header_row + 1 (1-based)
    header_row_1based = header_row + 1
    col_name_to_idx = {}

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row_1based, column=col_idx)
        if cell.value is not None:
            col_name = str(cell.value).strip()
            col_name_to_idx[col_name] = col_idx

    # Converter valores nas colunas selecionadas
    converted_cols = []
    errors = []
    data_start_row = header_row_1based + 1

    for col_name in columns_to_convert:
        if col_name not in col_name_to_idx:
            errors.append(f'Coluna "{col_name}" não encontrada')
            continue

        col_idx = col_name_to_idx[col_name]
        converted = False

        for row_idx in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            original_value = cell.value

            if original_value is None:
                continue

            numeric_val = clean_numeric_value(original_value)
            if numeric_val is not np.nan and not np.isnan(numeric_val):
                cell.value = numeric_val / divisor
                # Manter formato numérico se não tiver um
                if cell.number_format == 'General' or cell.number_format is None:
                    cell.number_format = '#,##0.00'
                converted = True

        if converted:
            converted_cols.append(col_name)
        else:
            errors.append(f'Nenhum valor numérico encontrado em "{col_name}"')

    wb.save(output_path)
    wb.close()

    return converted_cols, errors


def convert_csv_preserving(filepath, output_path, header_row, columns_to_convert, divisor):
    """Converte CSV - lê e grava mantendo todas as colunas originais."""
    df = pd.read_csv(filepath, sep=None, engine='python', header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    converted_cols = []
    errors = []

    for col in columns_to_convert:
        if col not in df.columns:
            errors.append(f'Coluna "{col}" não encontrada')
            continue
        try:
            df[col] = df[col].apply(clean_numeric_value) / divisor
            converted_cols.append(col)
        except Exception as e:
            errors.append(f'Erro em "{col}": {str(e)}')

    df.to_csv(output_path, index=False)
    return converted_cols, errors


# ─── Routes ───────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory(STATIC_FOLDER, 'index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.xlsx', '.xls', '.csv', '.ods']:
        return jsonify({'error': 'Formato não suportado. Use .xlsx, .xls, .csv ou .ods'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, 'current_file' + ext)
    file.save(filepath)

    try:
        if ext == '.csv':
            df_raw = pd.read_csv(filepath, sep=None, engine='python', header=None)
        else:
            df_raw = pd.read_excel(filepath, header=None)
    except Exception as e:
        return jsonify({'error': f'Erro ao ler arquivo: {str(e)}'}), 400

    header_row = detect_header_row(df_raw)

    try:
        if ext == '.csv':
            df = pd.read_csv(filepath, sep=None, engine='python', header=header_row)
        else:
            df = pd.read_excel(filepath, header=header_row)
    except Exception as e:
        return jsonify({'error': f'Erro ao ler arquivo: {str(e)}'}), 400

    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how='all').reset_index(drop=True)

    columns_info = analyze_columns(df)

    meta = {
        'filepath': filepath,
        'ext': ext,
        'filename': file.filename,
        'header_row': header_row,
    }
    with open(os.path.join(UPLOAD_FOLDER, 'meta.json'), 'w') as f:
        json.dump(meta, f)

    preview = sanitize_preview(df.head(10).fillna('').to_dict(orient='records'))

    return jsonify({
        'columns': df.columns.tolist(),
        'columns_info': columns_info,
        'row_count': len(df),
        'header_row': header_row,
        'preview': preview,
        'filename': file.filename
    })


@app.route('/api/convert', methods=['POST'])
def convert():
    data = request.get_json()
    columns_to_convert = data.get('columns', [])
    divisor = data.get('divisor', 1000)

    if not columns_to_convert:
        return jsonify({'error': 'Nenhuma coluna selecionada'}), 400

    try:
        with open(os.path.join(UPLOAD_FOLDER, 'meta.json'), 'r') as f:
            meta = json.load(f)
    except FileNotFoundError:
        return jsonify({'error': 'Nenhum arquivo carregado.'}), 400

    filepath = meta['filepath']
    ext = meta['ext']
    header_row = meta.get('header_row', 0)

    output_filename = 'convertido_' + meta['filename']

    if ext == '.csv':
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)
        converted_cols, errors = convert_csv_preserving(
            filepath, output_path, header_row, columns_to_convert, divisor
        )
    else:
        # Para Excel: preservar formatação completa
        if not output_filename.lower().endswith('.xlsx'):
            output_filename = os.path.splitext(output_filename)[0] + '.xlsx'

        # Se for .xls, primeiro converter para .xlsx preservando dados
        if ext == '.xls':
            # Converter .xls para .xlsx primeiro
            temp_xlsx = os.path.join(UPLOAD_FOLDER, 'temp_converted.xlsx')
            df_full = pd.read_excel(filepath, header=None)
            df_full.to_excel(temp_xlsx, index=False, header=False)
            source_path = temp_xlsx
        else:
            source_path = filepath

        output_path = os.path.join(UPLOAD_FOLDER, output_filename)

        # Copiar arquivo original para output (preserva tudo)
        import shutil
        shutil.copy2(source_path, output_path)

        # Converter apenas os valores nas colunas selecionadas
        converted_cols, errors = convert_excel_preserving_format(
            output_path, output_path, header_row, columns_to_convert, divisor
        )

    with open(os.path.join(UPLOAD_FOLDER, 'output_meta.json'), 'w') as f:
        json.dump({'output_path': output_path, 'output_filename': output_filename}, f)

    # Gerar preview dos dados convertidos
    try:
        if ext == '.csv':
            df_preview = pd.read_csv(output_path, sep=None, engine='python')
        else:
            df_preview = pd.read_excel(output_path, header=header_row)
        df_preview.columns = [str(c).strip() for c in df_preview.columns]
        df_preview = df_preview.dropna(how='all').reset_index(drop=True)
        preview = sanitize_preview(df_preview.head(10).fillna('').to_dict(orient='records'))
        columns = df_preview.columns.tolist()
    except:
        preview = []
        columns = []

    return jsonify({
        'success': True,
        'converted_columns': converted_cols,
        'errors': errors,
        'preview': preview,
        'columns': columns,
        'output_filename': output_filename,
        'row_count': len(df_preview) if 'df_preview' in dir() else 0
    })


@app.route('/api/download')
def download():
    try:
        with open(os.path.join(UPLOAD_FOLDER, 'output_meta.json'), 'r') as f:
            output_meta = json.load(f)
    except FileNotFoundError:
        return jsonify({'error': 'Nenhum arquivo convertido disponível'}), 404

    return send_file(
        output_meta['output_path'],
        as_attachment=True,
        download_name=output_meta['output_filename']
    )


if __name__ == '__main__':
    port = 5000
    print(f"\n{'='*50}")
    print(f"  Conversor Milheiro → Unidade")
    print(f"  Acesse: http://localhost:{port}")
    print(f"  Pressione Ctrl+C para encerrar")
    print(f"{'='*50}\n")

    # Abre o navegador automaticamente
    threading.Timer(1.5, lambda: webbrowser.open(f'http://localhost:{port}')).start()

    app.run(host='0.0.0.0', port=port, debug=False)